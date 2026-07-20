"""
Post-processing formatter for the LinkedIn-Schema-Tracker output.

Call format_excel(path) right after you save the DataFrame with
df.to_excel(path, index=False) — it will fix column widths, wrap
long text columns, freeze the header row, and make it look like a
real report instead of a raw dump.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Columns that should wrap and be wide
WIDE_TEXT_COLUMNS = {"Post Text", "Summary", "Remarks"}

# Sensible default widths per column name; anything not listed gets auto-fit
FIXED_WIDTHS = {
    "Sr. No.": 8,
    "Date Collected": 20,
    "Post Date": 12,
    "Author": 30,
    "Post URL": 25,
    "Category": 16,
    "Summary File": 40,
    "Status": 12,
}


def format_excel(path: str, sheet_name: str | None = None) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]

    # --- Header row styling ---
    header_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    thin_border = Border(left=Side(style='thin', color='FFD4D4D4'), 
                         right=Side(style='thin', color='FFD4D4D4'), 
                         top=Side(style='thin', color='FFD4D4D4'), 
                         bottom=Side(style='thin', color='FFD4D4D4'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = "A2"          # keep header visible while scrolling
    ws.auto_filter.ref = ws.dimensions  # add filter dropdowns

    # --- Column widths + wrapping ---
    for idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(idx)

        if header in WIDE_TEXT_COLUMNS:
            ws.column_dimensions[col_letter].width = 80
        elif header in FIXED_WIDTHS:
            ws.column_dimensions[col_letter].width = FIXED_WIDTHS[header]
        else:
            # auto-fit based on longest value in the column, capped at 30
            max_len = max(
                [len(str(header))]
                + [len(str(row[idx - 1].value)) for row in ws.iter_rows(min_row=2) if row[idx - 1].value]
                or [10]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # --- Row-level formatting: wrap text, top-align, row height, borders ---
    wide_col_indices = [i for i, h in enumerate(headers, start=1) if h in WIDE_TEXT_COLUMNS]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if cell.column in wide_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        # Allow Excel to naturally auto-fit row height for wrapped text
        # (Setting a fixed height causes text overlapping if the text is too long)

    # --- Alternating row shading for readability ---
    stripe_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = stripe_fill

    wb.save(path)


if __name__ == "__main__":
    import sys
    format_excel(sys.argv[1])
